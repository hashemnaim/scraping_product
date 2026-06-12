"""تصدير صفوف المنتجات إلى Excel."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pipeline.constants import EXCEL_COLUMNS, IMPORT_DEFAULTS


def build_row(
    product_id: int,
    product: dict,
    images_folder: str,
    module_id: int,
    category_id: int,
    sub_category_id: int,
    import_defaults: dict | None = None,
) -> dict:
    defaults = {**IMPORT_DEFAULTS, **(import_defaults or {})}
    image_path = product.get("image_path", "")
    if not image_path and product.get("image_ok"):
        from pipeline.images import image_relative_path

        image_path = image_relative_path(images_folder, product_id)

    tags = product.get("tags") or product.get("category", "") or defaults.get("Tags", "")

    return {
        "Id": product_id,
        "Name": product.get("name", ""),
        "Description": product.get("description", ""),
        "Image": image_path,
        "CategoryId": category_id,
        "SubCategoryId": sub_category_id,
        "BrandId": defaults.get("BrandId", ""),
        "UnitId": product.get("unit_id", defaults.get("UnitId", "")),
        "Stock": defaults.get("Stock", 100),
        "Price": product.get("price", ""),
        "Discount": defaults.get("Discount", 0),
        "DiscountType": defaults.get("DiscountType", "percent"),
        "AvailableTimeStarts": defaults.get("AvailableTimeStarts", ""),
        "AvailableTimeEnds": defaults.get("AvailableTimeEnds", ""),
        "Variations": defaults.get("Variations", ""),
        "ChoiceOptions": defaults.get("ChoiceOptions", ""),
        "AddOns": defaults.get("AddOns", ""),
        "Attributes": defaults.get("Attributes", ""),
        "Tags": tags,
        "StoreId": defaults.get("StoreId", ""),
        "ModuleId": module_id,
        "Status": defaults.get("Status", 1),
        "Veg": defaults.get("Veg", 0),
        "Recommended": defaults.get("Recommended", 0),
        "IsDefaultProduct": defaults.get("IsDefaultProduct", 1),
        "IsPrescriptionReq": defaults.get("IsPrescriptionReq", 0),
        "CommonConditions": defaults.get("CommonConditions", ""),
        "IsBasic": defaults.get("IsBasic", 1),
        "QuantityUnit": product.get("quantity_unit", defaults.get("QuantityUnit", "")),
    }


def write_excel(rows: list[dict], excel_path: Path) -> None:
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Products"

    header_fill = PatternFill("solid", start_color="1F4E79")
    alt_fill = PatternFill("solid", start_color="D6E4F0")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(EXCEL_COLUMNS, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    worksheet.row_dimensions[1].height = 30

    for row_index, row_data in enumerate(rows, 2):
        row_fill = alt_fill if row_index % 2 == 0 else None
        for col, column in enumerate(EXCEL_COLUMNS, 1):
            cell = worksheet.cell(row=row_index, column=col, value=row_data.get(column, ""))
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_fill:
                cell.fill = row_fill
        worksheet.row_dimensions[row_index].height = 25

    col_widths = {
        "Id": 14,
        "Name": 35,
        "Description": 40,
        "Image": 28,
        "CategoryId": 12,
        "SubCategoryId": 14,
        "Tags": 30,
        "ModuleId": 10,
        "Price": 12,
    }
    for col, header in enumerate(EXCEL_COLUMNS, 1):
        worksheet.column_dimensions[get_column_letter(col)].width = col_widths.get(header, 14)

    worksheet.freeze_panes = "A2"
    workbook.save(excel_path)
