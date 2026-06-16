"""تصدير محلات الخريطة إلى Excel."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pipeline.constants import PLACES_EXCEL_COLUMNS


def write_places_excel(rows: list[dict], excel_path: Path) -> None:
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Places"

    header_fill = PatternFill("solid", start_color="1F4E79")
    alt_fill = PatternFill("solid", start_color="D6E4F0")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(PLACES_EXCEL_COLUMNS, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    worksheet.row_dimensions[1].height = 30

    for row_index, row_data in enumerate(rows, 2):
        row_fill = alt_fill if row_index % 2 == 0 else None
        for col, column in enumerate(PLACES_EXCEL_COLUMNS, 1):
            cell = worksheet.cell(row=row_index, column=col, value=row_data.get(column, ""))
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_fill:
                cell.fill = row_fill
        worksheet.row_dimensions[row_index].height = 22

    widths = {
        "Id": 8,
        "Name": 32,
        "Address": 40,
        "Phone": 18,
        "Category": 18,
        "City": 16,
        "CoverageKm": 12,
        "MapsUrl": 36,
    }
    for col, header in enumerate(PLACES_EXCEL_COLUMNS, 1):
        worksheet.column_dimensions[get_column_letter(col)].width = widths.get(header, 14)

    worksheet.freeze_panes = "A2"
    workbook.save(excel_path)
