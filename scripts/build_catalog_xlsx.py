"""إنشاء ملفات Excel للكتالوج من catalog/seeds/admin_catalog.json."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
CATALOG = ROOT / "catalog"
SEED_PATH = CATALOG / "seeds" / "admin_catalog.json"


def _header_row(sheet, headers):
    fill = PatternFill("solid", start_color="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.fill = fill
        cell.font = font


def _slug_or_default(slug: str | None, item_id: int) -> str:
    if slug and str(slug).strip().lower() not in {"", "null", "none"}:
        return str(slug).strip()
    return f"sub-{item_id}"


def load_seed() -> dict:
    return json.loads(SEED_PATH.read_text(encoding="utf-8"))


def build_modules(data: dict) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modules"
    _header_row(ws, ["ModuleId", "NameAr", "Active", "Stock", "Status", "IsDefaultProduct", "IsBasic"])
    for row, item in enumerate(data["modules"], 2):
        active = 1 if int(item.get("status", 0)) == 1 else 0
        ws.cell(row, 1, item["id"])
        ws.cell(row, 2, item.get("module_name", ""))
        ws.cell(row, 3, active)
        ws.cell(row, 4, 100)
        ws.cell(row, 5, 1)
        ws.cell(row, 6, 1)
        ws.cell(row, 7, 1)
    wb.save(CATALOG / "modules.xlsx")


def build_categories_and_subs(data: dict) -> None:
    wb_flat = openpyxl.Workbook()
    ws_flat = wb_flat.active
    ws_flat.title = "CategoriesFlat"
    _header_row(ws_flat, ["id", "name", "parent_id", "module_id", "slug"])

    wb_cat = openpyxl.Workbook()
    ws_cat = wb_cat.active
    ws_cat.title = "Categories"
    _header_row(ws_cat, ["ModuleId", "CategoryId", "NameAr"])

    wb_sub = openpyxl.Workbook()
    ws_sub = wb_sub.active
    ws_sub.title = "SubCategories"
    _header_row(
        ws_sub,
        [
            "ModuleId",
            "CategoryId",
            "SubCategoryId",
            "NameAr",
            "DefaultSourceUrl",
            "OutputSlug",
            "ExcelFilename",
            "ImagesFolder",
        ],
    )

    cat_row = 2
    sub_row = 2
    flat_row = 2

    for item in data["categories"]:
        module_id = int(item["module_id"])
        item_id = int(item["id"])
        name = item.get("name", "")
        slug = item.get("slug")

        ws_flat.cell(flat_row, 1, item_id)
        ws_flat.cell(flat_row, 2, name)
        ws_flat.cell(flat_row, 3, 0)
        ws_flat.cell(flat_row, 4, module_id)
        ws_flat.cell(flat_row, 5, slug or "")
        flat_row += 1

        ws_cat.cell(cat_row, 1, module_id)
        ws_cat.cell(cat_row, 2, item_id)
        ws_cat.cell(cat_row, 3, name)
        cat_row += 1

    for item in data.get("subcategories", []):
        module_id = int(item["module_id"])
        item_id = int(item["id"])
        parent_id = int(item.get("parent_id") or 0)
        name = item.get("name", "")
        slug = _slug_or_default(item.get("slug"), item_id)

        ws_flat.cell(flat_row, 1, item_id)
        ws_flat.cell(flat_row, 2, name)
        ws_flat.cell(flat_row, 3, parent_id)
        ws_flat.cell(flat_row, 4, module_id)
        ws_flat.cell(flat_row, 5, slug)
        flat_row += 1

        ws_sub.cell(sub_row, 1, module_id)
        ws_sub.cell(sub_row, 2, parent_id)
        ws_sub.cell(sub_row, 3, item_id)
        ws_sub.cell(sub_row, 4, name)
        ws_sub.cell(sub_row, 5, "")
        ws_sub.cell(sub_row, 6, slug)
        ws_sub.cell(sub_row, 7, f"{slug}.xlsx")
        ws_sub.cell(sub_row, 8, "product_images")
        sub_row += 1

    wb_cat.save(CATALOG / "categories.xlsx")
    wb_sub.save(CATALOG / "subcategories.xlsx")
    wb_flat.save(CATALOG / "categories_flat.xlsx")


def build_units(data: dict) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Units"
    _header_row(ws, ["ModuleId", "UnitId", "NameAr", "Aliases"])
    for row, item in enumerate(data["units"], 2):
        ws.cell(row, 1, item["module_id"])
        ws.cell(row, 2, item["id"])
        ws.cell(row, 3, item.get("unit", ""))
        ws.cell(row, 4, "")
    wb.save(CATALOG / "units.xlsx")


if __name__ == "__main__":
    seed = load_seed()
    CATALOG.mkdir(parents=True, exist_ok=True)
    build_modules(seed)
    build_categories_and_subs(seed)
    build_units(seed)
    print("Created catalog/*.xlsx from catalog/seeds/admin_catalog.json")
