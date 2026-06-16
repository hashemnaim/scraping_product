"""اختبارات قراءة صيغة تصدير لوحة الإدارة."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from pipeline import catalog


@pytest.fixture()
def catalog_dir(tmp_path, monkeypatch):
    cat = tmp_path / "catalog"
    cat.mkdir()
    monkeypatch.setattr(catalog, "_catalog_dir", lambda: cat)
    monkeypatch.setattr(catalog, "_modules_xlsx", lambda: cat / "modules.xlsx")
    monkeypatch.setattr(catalog, "_categories_xlsx", lambda: cat / "categories.xlsx")
    monkeypatch.setattr(catalog, "_subcategories_xlsx", lambda: cat / "subcategories.xlsx")
    monkeypatch.setattr(catalog, "_units_xlsx", lambda: cat / "units.xlsx")
    catalog.clear_cache()
    return cat


def _write_modules(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "module_name", "status"])
    ws.append([3, "سوبر ماركت", 1])
    wb.save(path)


def _write_flat_categories(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "name", "parent_id", "module_id", "slug"])
    ws.append([243, "الفواكه والخضروات", 0, 3, None])
    ws.append([216, "فواكه", 243, 3, "foakh216"])
    wb.save(path)


def _write_admin_units(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "module_id", "unit"])
    ws.append([9, 3, "كيلو"])
    ws.append([10, 3, "غرام"])
    wb.save(path)


def test_flat_category_export_splits_by_parent_id(catalog_dir):
    _write_modules(catalog_dir / "modules.xlsx")
    _write_flat_categories(catalog_dir / "categories.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "name", "parent_id", "module_id", "slug"])
    ws.append([216, "فواكه", 243, 3, "foakh216"])
    wb.save(catalog_dir / "subcategories.xlsx")
    catalog.clear_cache()

    categories = catalog.list_categories(3)
    assert len(categories) == 1
    assert categories[0].category_id == 243

    subs = catalog.list_subcategories(3, 243)
    assert len(subs) == 1
    assert subs[0].sub_category_id == 216
    assert subs[0].output_slug == "foakh216"


def test_admin_units_respect_module_id(catalog_dir):
    _write_modules(catalog_dir / "modules.xlsx")
    _write_admin_units(catalog_dir / "units.xlsx")
    catalog.clear_cache()

    units = catalog.get_units(3)
    assert len(units) == 2
    assert catalog.get_units(1) == []


def test_module_id_zero_inherits_from_parent(catalog_dir, monkeypatch):
    seed_dir = catalog_dir / "seeds"
    seed_dir.mkdir()
    seed_dir.joinpath("admin_catalog.json").write_text(
        """{
          "modules": [{"id": 3, "module_name": "سوبر ماركت", "status": 1}],
          "units": [],
          "categories": [
            {"id": 243, "name": "الفواكه والخضروات", "module_id": 3, "slug": null}
          ],
          "subcategories": [
            {"id": 216, "name": "فواكه", "parent_id": 243, "module_id": 0, "slug": "foakh216"}
          ]
        }""",
        encoding="utf-8",
    )
    _write_modules(catalog_dir / "modules.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "name", "parent_id", "module_id", "slug"])
    ws.append([243, "الفواكه والخضروات", 0, 3, None])
    wb.save(catalog_dir / "categories.xlsx")

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.append(["id", "name", "parent_id", "module_id", "slug"])
    ws2.append([216, "فواكه", 243, 0, "foakh216"])
    wb2.save(catalog_dir / "subcategories.xlsx")
    catalog.clear_cache()

    subs = catalog.list_subcategories(3, 243)
    assert len(subs) == 1
    assert subs[0].sub_category_id == 216
    assert subs[0].module_id == 3
