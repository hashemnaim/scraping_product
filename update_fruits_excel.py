"""تحديث fruits.xlsx: أسماء عربية، CategoryId، UnitId، QuantityUnit."""

import re
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(__file__).parent / "fruits-vegetables" / "fruits.xlsx"
CATEGORY_ID = 253

ARABIC_NAMES = {
    "Spanish Cantaloup": "شمام إسباني",
    "Prepared Pineapple": "أناناس مقطع",
    "Imported Pineapple": "أناناس مستورد",
    "Cherry": "كرز",
    "Seoudi Prepared Yellow Melon (Almas)": "شمام أصفر مقطع سعودي (ألماس)",
    "Orange for Juice": "برتقال للعصير",
    "Pico Blackberry": "توت أسود بيكو",
    "Seoudi Prepared Watermelon": "بطيخ مقطع سعودي",
    "Seoudi Prepared Strawberry": "فراولة مقطعة سعودي",
    "Seoudi Prepared Kiwi": "كيوي مقطع سعودي",
    "Seoudi Prepared Honeyglow Pineapple": "أناناس هوني جلو مقطع سعودي",
    "Seoudi Prepared Cantaloupe": "شمام مقطع سعودي",
    "Pico Bashmala": "باشمالا بيكو",
    "Belco Bluebery": "توت أزرق بيلكو",
    "Yellow Medium Apple": "تفاح أصفر متوسط",
    "Belco Black Berry": "توت أسود بيلكو",
    "Sukkary Apple 65/70": "تفاح سكري 65/70",
    "Medium Red Apples": "تفاح أحمر متوسط",
    "Greek Figs": "تين يوناني",
    "Seoudi Fruit Salad": "سلطة فواكه سعودي",
    "Seoudi Prepared Cantaloup": "شمام مقطع سعودي",
    "Longan": "لونجان",
    "New Zealand Yellow Kiwi": "كيوي أصفر نيوزيلندي",
    "Belco Raspberry": "توت العليق بيلكو",
    "Pico Raspberry": "توت العليق بيكو",
    "Siafa Mabroom Dates": "تمور مبروم صيافة",
    "Muskmelon": "شمام",
    "Large Majdool Dates": "تمور مجدول كبيرة",
    "Imported Flat Peaches": "خوخ مسطح مستورد",
    "Yellow Melon (AlMas)": "شمام أصفر (ألماس)",
    "Egyptian Black Grapes": "عنب أسود مصري",
    "Crop's Blackcurrants": "كشمش أسود كروبس",
    "Hygiene Dutch Blueberries": "توت أزرق هولندي هيجين",
    "Dutch Black Raspberries": "توت العليق أسود هولندي",
    "Hygiene Dutch Red Berries": "توت أحمر هولندي هيجين",
    "Siafa Safawi Dates": "تمور صفاوي صيافة",
    "Eva Grow Omani White Mulberries": "توت أبيض عماني إيفا جرو",
    "Mafaza Blueberries": "توت أزرق مافازا",
    "Pico Blueberry": "توت أزرق بيكو",
    "Imported White Dragon Fruit": "فاكهة التنين البيضاء مستوردة",
    "Rambutan": "رامبوتان",
    "Mangosteen": "مانجستين",
    "Siafa Sagai Dates": "تمور سكري صيافة",
    "Daltex Valencia Oranges": "برتقال فالنسيا دالتكس",
    "Ambrosia Apple": "تفاح أمبروسيا",
    "Siafa Sukkary Mefatel Dates": "تمور سكري مفتل صيافة",
    "Baladi Apricots": "مشمش بلدي",
    "Baladi Peaches": "خوخ بلدي",
    "Sweet Peaches": "خوخ حلو",
    "Egyptian Red Grapes": "عنب أحمر مصري",
    "Egyptian White Grapes": "عنب أبيض مصري",
    "Sweet Plum": "برقوق حلو",
    "Hollywoods Plum": "برقوق هوليوود",
    "Pumpkin": "يقطين",
    "Imported Cherries": "كرز مستورد",
    "Imported Apricots": "مشمش مستورد",
    "Oranges for juice": "برتقال للعصير",
    "Mafa Baladi Valencia Oranges": "برتقال فالنسيا بلدي مافا",
    "Mixed Peppers": "فلفل مشكل",
    "Potatoes": "بطاطس",
    "Red Onions": "بصل أحمر",
    "Omani Red Berries": "توت أحمر عماني",
    "Nature's Pick Blueberries": "توت أزرق نيتشرز بيك",
    "Nature’s Pick Blueberries": "توت أزرق نيتشرز بيك",
    "Blueberries": "توت أزرق",
    "Prepared Imported Sukkary Pineapple": "أناناس سكري مقطع مستورد",
    "Crop's Cranberries": "توت بري كروبس",
    "Baby Pumpkin": "يقطين صغير",
    "Almas Slices": "شرائح ألماس",
    "Seoudi Prepared Dragon Fruit": "فاكهة التنين مقطعة سعودي",
    "Zucchini": "كوسة",
    "Mafa Strawberry": "فراولة مافا",
    "Nature's Pick Blueberry": "توت أزرق نيتشرز بيك",
    "Nature’s Pick Blueberry": "توت أزرق نيتشرز بيك",
    "Nature's Pick Blue Berry": "توت أزرق نيتشرز بيك",
    "Nature’s Pick Blue Berry": "توت أزرق نيتشرز بيك",
    "Belco Blueberry": "توت أزرق بيلكو",
    "Hygine Local Blue Berry": "توت أزرق محلي هيجين",
    "Seoudi Prepared Coconut": "جوز هند مقطع سعودي",
    "Seoudi Prepared Oranges": "برتقال مقطع سعودي",
    "Seoudi Prepared Groundcherry": "فاكهة الترمس الذهبي مقطعة سعودي",
    "Seoudi Prepared Red Grapes": "عنب أحمر مقطع سعودي",
    "Seoudi Prepared Orange": "برتقال مقطع سعودي",
    "Natures Bake Raspberry": "توت العليق نيتشرز بيك",
    "Nature's Pick Blackberry": "توت أسود نيتشرز بيك",
    "Nature’s Pick Blackberry": "توت أسود نيتشرز بيك",
    "Prepared Jackfruit": "جاك فروت مقطع",
}

# استخراج الوحدة من الاسم الإنجليزي الأصلي (حسب SKU)
SKU_UNITS = {
    "6001254042444": ("250", 3),
    "154938": ("200", 3),
    "155039": ("2", 2),
    "6223001530121": ("125", 3),
    "154906": ("300", 3),
    "154911": ("200", 3),
    "154913": ("200", 3),
    "154903": ("200", 3),
    "154912": ("225", 3),
    "6223001530930": ("500", 3),
    "6222037090975": ("125", 3),
    "6222010801161": ("150", 3),
    "153204": ("250", 3),
    "154008": ("350", 3),
    "6222010817360": ("125", 3),
    "6223001531210": ("125", 3),
    "5410355300970": ("1", 2),
    "761635202503": ("125", 3),
    "761635203401": ("125", 3),
    "133144": ("125", 3),
    "6223004379581": ("200", 3),
    "6224009117192": ("125", 3),
    "6223001530114": ("125", 3),
    "6223002335398": ("2", 2),
    "6223000250112": ("2", 2),
    "149414": ("500", 3),
    "149412": ("1", 2),
    "133038": ("1", 2),
    "0761742517156": ("500", 3),
    "6009703401538": ("125", 3),
    "5410355305692": ("1", 2),
    "160380": ("200", 3),
    "157569": ("1", 2),
    "6223000253823": ("1", 2),
    "5411683221050": ("250", 3),
    "3800040800328": ("125", 3),
    "6222037091439": ("1", 2),
    "6221034098799": ("125", 3),
    "155392": ("200", 3),
    "155346": ("200", 3),
    "154907": ("200", 3),
    "154910": ("200", 3),
    "154904": ("400", 3),
    "0710535263138": ("125", 3),
    "3800228450086": ("125", 3),
    "6222037090753": ("250", 3),
    "151590": ("250", 3),
}


STOP_WORDS = {
    "مقطع",
    "مقطعة",
    "مقطوف",
    "مستورد",
    "مستوردة",
    "للعصير",
    "من",
    "و",
    "صغير",
    "كبيرة",
    "متوسط",
    "هولندي",
    "هيجين",
    "نيوزيلندي",
    "يوناني",
    "عماني",
    "هوليوود",
}

FRUIT_PHRASES = [
    "فاكهة الترمس الذهبي",
    "فاكهة التنين",
    "سلطة فواكه",
    "توت العليق",
    "توت أزرق",
    "توت أسود",
    "توت أحمر",
    "توت أبيض",
    "توت بري",
    "جوز هند",
    "جاك فروت",
    "شمام أصفر",
    "تفاح أصفر",
    "تفاح أحمر",
    "تفاح سكري",
    "عنب أسود",
    "عنب أحمر",
    "عنب أبيض",
    "برتقال فالنسيا",
    "برتقال للعصير",
    "خوخ مسطح",
    "كشمش أسود",
    "تمور مجدول",
    "تمور مبروم",
    "تمور صفاوي",
    "تمور سكري",
    "أناناس هوني جلو",
    "أناناس سكري",
]

FRUIT_SYNONYMS = {
    "شمام": ["شمام", "الشمام", "شمامات", "بطيخ أصفر", "كانتلوب", "كنتالوب", "شمام طازج"],
    "أناناس": ["أناناس", "الأناناس", "اناناس", "أناناسة", "اناناسة"],
    "كرز": ["كرز", "الكرز", "كرزات", "كرزة", "كرز طازج"],
    "برتقال": ["برتقال", "البرتقال", "برتقالة", "برتقالات", "برتقال طازج"],
    "بطيخ": ["بطيخ", "البطيخ", "بطيخة", "بطيخ طازج"],
    "فراولة": ["فراولة", "الفراولة", "فراول", "فراولة طازجة"],
    "كيوي": ["كيوي", "الكيوي", "كيوي طازج"],
    "توت": ["توت", "التوت", "توتات", "توت طازج"],
    "تفاح": ["تفاح", "التفاح", "تفاحة", "تفاحات", "تفاح طازج"],
    "تين": ["تين", "التين", "تينة", "تين طازج"],
    "لونجان": ["لونجان", "لونجام", "عين التنين"],
    "تمور": ["تمور", "التمر", "تمر", "رطب", "بلح", "تمور طازجة"],
    "خوخ": ["خوخ", "الخوخ", "خوخة", "خوخ طازج"],
    "عنب": ["عنب", "العنب", "عنبة", "عناقيد", "عنب طازج"],
    "برقوق": ["برقوق", "البرقوق", "برقوقة"],
    "مشمش": ["مشمش", "المشمش", "مشمشة", "مشمش طازج"],
    "يقطين": ["يقطين", "اليقطين", "قرع", "يقطين طازج"],
    "كشمش": ["كشمش", "الكشمش", "زبيب", "عنب مجفف"],
    "مانجستين": ["مانجستين", "مانجا ستين"],
    "رامبوتان": ["رامبوتان", "رامبوتان طازج"],
    "باشمالا": ["باشمالا", "بشملة", "توت بري"],
    "فلفل": ["فلفل", "الفلفل", "فلفل رومي", "فلفل ملون"],
    "بطاطس": ["بطاطس", "البطاطس", "بطاطا"],
    "بصل": ["بصل", "البصل", "بصلة"],
    "كوسة": ["كوسة", "الكوسة", "كوسا", "زوكيني"],
    "موز": ["موز", "الموز", "موزة", "موز طازج"],
}

MODIFIER_TAGS = {
    "بلدي": ["بلدي", "بلدية", "محلي", "محلية"],
    "مصري": ["مصري", "مصرية", "من مصر"],
    "إسباني": ["إسباني", "إسبانية", "من إسبانيا"],
    "مستورد": ["مستورد", "مستوردة", "مستورده", "مستوردات"],
    "مقطع": ["مقطع", "مقطعة", "مقطوف", "جاهز", "مفرز"],
    "سعودي": ["سعودي", "سعودية", "منتجات سعودي", "سعودى"],
    "أصفر": ["أصفر", "صفراء", "أصفره"],
    "أحمر": ["أحمر", "حمراء", "حمره"],
    "أسود": ["أسود", "سوداء", "سوده"],
    "أبيض": ["أبيض", "بيضاء", "بيضه"],
    "أزرق": ["أزرق", "زرقاء"],
    "حلو": ["حلو", "حلوة", "حلوه"],
    "طازج": ["طازج", "طازجة", "طازه"],
    "عصير": ["عصير", "عصائر", "للعصير"],
}

BRAND_TAGS = {
    "سعودي": ["سعودي", "منتجات سعودي"],
    "بيكو": ["بيكو", "pico"],
    "بيلكو": ["بيلكو", "belco"],
    "صيافة": ["صيافة", "سيافة"],
    "دالتكس": ["دالتكس", "daltex"],
    "مافا": ["مافا", "mafa"],
    "مافازا": ["مافازا", "mafaza"],
    "نيتشرز": ["نيتشرز بيك", "natures pick", "نيتشر"],
    "كروبس": ["كروبس", "crops"],
    "هيجين": ["هيجين", "hygiene"],
    "ألماس": ["ألماس", "almas"],
}


def with_al(word):
    word = word.strip()
    if len(word) < 2 or word.startswith("ال"):
        return None
    return f"ال{word}"


def build_search_tags(name):
    """بناء كلمات مفتاحية كاملة للبحث من اسم المنتج."""
    name = normalize_name(name)
    tags = []

    def add(*values):
        for value in values:
            if not value:
                continue
            value = str(value).strip()
            if value and value not in tags:
                tags.append(value)

    add(name, "فواكه", "فاكهة", "فواكه طازجة", "فاكهة طازجة")

    for phrase in sorted(FRUIT_PHRASES, key=len, reverse=True):
        if phrase in name:
            add(phrase, with_al(phrase.split()[0]))

    for root, synonyms in sorted(FRUIT_SYNONYMS.items(), key=lambda item: -len(item[0])):
        if root in name:
            add(root, with_al(root))
            add(*synonyms)
            break

    tokens = [
        token
        for token in re.split(r"[\s()\-]+", name)
        if token and token not in STOP_WORDS and len(token) > 1 and not token.isdigit()
    ]

    for index in range(len(tokens)):
        for length in range(2, min(5, len(tokens) - index + 1)):
            add(" ".join(tokens[index : index + length]))

    for token in tokens:
        add(token, with_al(token))

    for key, extras in MODIFIER_TAGS.items():
        if key in name:
            add(key, *extras)

    for key, extras in BRAND_TAGS.items():
        if key in name:
            add(key, *extras)

    if "للعصير" in name or "عصير" in name:
        add("برتقال عصير", "فواكه عصير", "عصائر")

    if "سلطة" in name:
        add("سلطة", "سلطة فواكه", "فواكه مشكلة", "فاكهة مشكلة")

    if "مجدول" in name or "مبروم" in name or "صفاوي" in name or "سكري" in name:
        add("تمر", "تمور", "رطب", "بلح")

    return ", ".join(tags)


def normalize_name(name):
    return (name or "").replace("’", "'").strip()


def parse_unit(name):
    name = normalize_name(name)
    patterns = [
        (r"[-–]\s*(\d+(?:[.,]\d+)?)\s*(kg|kilo|kilogram)\b", 2),
        (r"[-–]\s*(\d+(?:[.,]\d+)?)\s*(g|gram|grams|gr)\b", 3),
        (r"(\d+(?:[.,]\d+)?)\s*(kg|kilo|kilogram)\b", 2),
        (r"(\d+(?:[.,]\d+)?)\s*(g|gram|grams|gr)\b", 3),
        (r"[-–]\s*(\d+)\s*Piece\b", None),
    ]
    for pattern, unit_id in patterns:
        match = re.search(pattern, name, re.I)
        if match:
            quantity = match.group(1).replace(",", ".")
            if quantity.endswith(".0"):
                quantity = quantity[:-2]
            clean_name = re.sub(pattern, "", name, flags=re.I).strip(" -–")
            if unit_id is None:
                return clean_name, None, None
            return clean_name, quantity, unit_id
    return name, None, None


def to_arabic_name(clean_name):
    clean_name = normalize_name(clean_name)
    return ARABIC_NAMES.get(clean_name, clean_name)


ORDERED_SKUS = [
    "2005700000000", "2005699000000", "2005698000000", "6001254042444", "154938",
    "155039", "6223001530121", "154906", "154911", "154913", "154903", "154912",
    "6223001530930", "6222037090975", "2005668000000", "6222010801161", "2005656000000",
    "2005135000000", "153204", "151732", "154008", "2000142000000", "2005039000000",
    "6222010817360", "6223001531210", "2000378000000", "2005630000000", "2005033000000",
    "2000318000000", "2008230000000", "2005008000000", "5410355300970", "761635202503",
    "761635203401", "133144", "2005136000000", "6223004379581", "6224009117192",
    "6223001530114", "2000074000000", "2000496000000", "2000495000000", "2000482000000",
    "6223002335398", "2000664000000", "2000433000000", "2000424000000", "2000623000000",
    "2000624000000", "2000629000000", "2000630000000", "2005003000000", "2005004000000",
    "2000649000000", "2000662000000", "2000670000000", "2000676000000", "2000683000000",
    "6223000250112", "149414", "149412", "133038", "2005795000000", "0761742517156",
    "6009703401538", "2005776000000", "5410355305692", "2005030000000", "2005742000000",
    "160380", "157569", "6223000253823", "5411683221050", "3800040800328", "6222037091439",
    "6221034098799", "155392", "155346", "154907", "154910", "154904", "0710535263138",
    "3800228450086", "6222037090753", "151590",
]

IMAGES_FOLDER_NAME = "fruits_images"
IMAGE_ID_DIGITS = 3


def sync_id_and_images(worksheet, columns):
    """مطابقة Id مع رقم الصورة: Id=1 -> product_001.webp"""
    images_dir = EXCEL_PATH.parent / IMAGES_FOLDER_NAME
    if not images_dir.is_dir():
        return

    for row in range(2, worksheet.max_row + 1):
        seq_id = row - 1
        new_name = f"product_{seq_id:0{IMAGE_ID_DIGITS}d}.webp"
        new_path = f"{IMAGES_FOLDER_NAME}/{new_name}"

        old_image = worksheet.cell(row, columns["Image"]).value or ""
        old_name = old_image.split("/")[-1] if old_image else ""
        old_file = images_dir / old_name
        new_file = images_dir / new_name

        if old_name and old_name != new_name and old_file.exists():
            if new_file.exists():
                new_file.unlink()
            old_file.rename(new_file)

        worksheet.cell(row, columns["Id"]).value = seq_id
        worksheet.cell(row, columns["Image"]).value = new_path


def main():
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    worksheet = workbook["Products"]
    columns = {worksheet.cell(1, col).value: col for col in range(1, worksheet.max_column + 1)}

    sync_id_and_images(worksheet, columns)

    for row in range(2, worksheet.max_row + 1):
        seq_id = row - 1
        sku = ORDERED_SKUS[seq_id - 1] if seq_id - 1 < len(ORDERED_SKUS) else ""
        arabic_name = worksheet.cell(row, columns["Name"]).value
        quantity, unit_id = SKU_UNITS.get(sku, (None, None))

        worksheet.cell(row, columns["Id"]).value = seq_id
        worksheet.cell(row, columns["Name"]).value = arabic_name
        worksheet.cell(row, columns["CategoryId"]).value = CATEGORY_ID
        worksheet.cell(row, columns["Tags"]).value = build_search_tags(arabic_name)
        worksheet.cell(row, columns["UnitId"]).value = unit_id
        worksheet.cell(row, columns["QuantityUnit"]).value = quantity
        worksheet.cell(row, columns["ModuleId"]).value = 1
        worksheet.cell(row, columns["IsDefaultProduct"]).value = 1
        worksheet.cell(row, columns["Image"]).value = (
            f"{IMAGES_FOLDER_NAME}/product_{seq_id:0{IMAGE_ID_DIGITS}d}.webp"
        )

    workbook.save(EXCEL_PATH)
    print(f"Updated {worksheet.max_row - 1} rows in {EXCEL_PATH}")


if __name__ == "__main__":
    main()
